#!/usr/bin/env python3
"""
generate_excel_templates.py
Generates sample Excel template files for bulk upload of master data.
Creates templates for: Vendors, Outlets, Categories, SubCategories, Makes, Models, Components
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Create templates directory
templates_dir = Path(__file__).parent.parent / "templates" / "excel"
templates_dir.mkdir(parents=True, exist_ok=True)

def create_header_row(ws, headers, required_cols=None):
    """Create a styled header row"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Mark required columns
        if required_cols and col_idx in required_cols:
            cell.value = f"{header} *"
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

def create_vendors_template():
    """Create Vendors Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendors"
    
    headers = ["Vendor Name", "Contact Person", "Email", "Mobile", "Address"]
    create_header_row(ws, headers, required_cols=[1])
    
    # Add sample data
    sample_data = [
        ["ABC Suppliers", "John Doe", "john@abc.com", "1234567890", "123 Main St"],
        ["XYZ Corporation", "Jane Smith", "jane@xyz.com", "9876543210", "456 Oak Ave"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "vendors_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

def create_outlets_template():
    """Create Outlets Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outlets"
    
    headers = ["Outlet Name", "Outlet Address", "Contact Info"]
    create_header_row(ws, headers, required_cols=[1])
    
    sample_data = [
        ["Amazon Online", "Online Portal", "support@amazon.in"],
        ["Croma Store", "Khar West Store", "022-11112222"],
        ["Reliance Digital", "Andheri East", "022-44443333"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "outlets_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

def create_categories_template():
    """Create Categories Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categories"
    
    headers = ["Category Name", "Description"]
    create_header_row(ws, headers, required_cols=[1])
    
    sample_data = [
        ["Electronics", "Electronic devices and gadgets"],
        ["Home Appliances", "Household appliances"],
        ["Smart Home Devices", "IoT and smart home products"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "categories_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

def create_subcategories_template():
    """Create SubCategories Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SubCategories"
    
    headers = ["SubCategory Name", "Category Name", "Description"]
    create_header_row(ws, headers, required_cols=[1])
    
    sample_data = [
        ["Smartphones", "Electronics", "Mobile phones"],
        ["Smart TVs", "Electronics", "Television sets"],
        ["Refrigerators", "Home Appliances", "Cooling appliances"],
        ["Washing Machines", "Home Appliances", "Laundry appliances"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "subcategories_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

def create_makes_template():
    """Create Makes Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Makes"
    
    headers = ["Make Name", "SubCategory Name"]
    create_header_row(ws, headers, required_cols=[1])
    
    sample_data = [
        ["Samsung", "Smart TVs"],
        ["LG", "Refrigerators"],
        ["Apple", "Smartphones"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "makes_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

def create_models_template():
    """Create Models Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Models"
    
    headers = ["Model Name", "Make Name", "Description"]
    create_header_row(ws, headers, required_cols=[1])
    
    sample_data = [
        ["iPhone 15 Pro", "Apple", "Latest iPhone model"],
        ["Samsung QLED 65", "Samsung", "65-inch QLED TV"],
        ["LG InstaView 260L", "LG", "260L refrigerator"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "models_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

def create_components_template():
    """Create Components Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Components"
    
    headers = ["Component Name", "Description"]
    create_header_row(ws, headers, required_cols=[1])
    
    sample_data = [
        ["Battery Pack", "Device rechargeable battery unit"],
        ["Charger", "Device adapter or charging cable"],
        ["Remote Control", "TV or AC remote controller"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file_path = templates_dir / "components_template.xlsx"
    wb.save(file_path)
    print(f"✅ Created: {file_path}")

if __name__ == "__main__":
    print("📊 Generating Excel templates for bulk upload...")
    print(f"📁 Output directory: {templates_dir}\n")
    
    try:
        create_vendors_template()
        create_outlets_template()
        create_categories_template()
        create_subcategories_template()
        create_makes_template()
        create_models_template()
        create_components_template()
        
        print(f"\n✅ All templates generated successfully in: {templates_dir}")
        print("\n📝 Note: These templates include sample data. Remove sample rows before uploading.")
    except Exception as e:
        print(f"❌ Error generating templates: {e}")
        import traceback
        traceback.print_exc()
