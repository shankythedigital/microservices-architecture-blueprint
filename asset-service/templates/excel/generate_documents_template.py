#!/usr/bin/env python3
"""
Generate Excel template for Documents Bulk Upload

This script creates an Excel template file for bulk uploading documents.
The template includes sample data and proper column headers.

Usage:
    python generate_documents_template.py

Output:
    documents_bulk_upload_template.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_excel_template():
    """Create Excel template with headers and sample data"""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documents"
    
    # Define headers
    headers = [
        "entity_type",
        "entity_id",
        "file_name",
        "file_path",
        "doc_type"
    ]
    
    # Define sample data
    sample_data = [
        ["ASSET", 1, "invoice_001.pdf", "/uploads/documents/invoice_001.pdf", "PDF"],
        ["ASSET", 2, "receipt_002.jpg", "/uploads/documents/receipt_002.jpg", "IMAGE"],
        ["COMPONENT", 1, "manual_001.pdf", "/uploads/documents/manual_001.pdf", "PDF"],
        ["WARRANTY", 1, "warranty_001.pdf", "/uploads/documents/warranty_001.pdf", "PDF"],
        ["AMC", 1, "amc_agreement_001.pdf", "/uploads/documents/amc_agreement_001.pdf", "AGREEMENT"],
        ["CATEGORY", 1, "category_image_001.jpg", "/uploads/documents/category_image_001.jpg", "IMAGE"],
        ["SUBCATEGORY", 1, "subcategory_image_001.jpg", "/uploads/documents/subcategory_image_001.jpg", "IMAGE"],
        ["MAKE", 1, "make_logo_001.png", "/uploads/documents/make_logo_001.png", "IMAGE"],
        ["MODEL", 1, "model_spec_001.pdf", "/uploads/documents/model_spec_001.pdf", "PDF"],
        ["OUTLET", 1, "outlet_photo_001.jpg", "/uploads/documents/outlet_photo_001.jpg", "IMAGE"],
        ["VENDOR", 1, "vendor_contract_001.pdf", "/uploads/documents/vendor_contract_001.pdf", "PDF"]
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write sample data
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx == 2:  # entity_id column - number format
                cell.number_format = "0"
    
    # Adjust column widths
    column_widths = {
        "A": 15,  # entity_type
        "B": 12,  # entity_id
        "C": 25,  # file_name
        "D": 50,  # file_path
        "E": 15   # doc_type
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add data validation for entity_type
    from openpyxl.worksheet.datavalidation import DataValidation
    
    entity_types = ["ASSET", "COMPONENT", "AMC", "WARRANTY", "CATEGORY", 
                    "SUBCATEGORY", "MAKE", "MODEL", "OUTLET", "VENDOR"]
    entity_type_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(entity_types)}"',
        allow_blank=True
    )
    entity_type_validation.add("A2:A1000")  # Apply to data rows
    ws.add_data_validation(entity_type_validation)
    
    # Add data validation for doc_type
    doc_types = ["PDF", "IMAGE", "RECEIPT", "AGREEMENT"]
    doc_type_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(doc_types)}"',
        allow_blank=True
    )
    doc_type_validation.add("E2:E1000")  # Apply to data rows
    ws.add_data_validation(doc_type_validation)
    
    # Save workbook
    filename = "documents_bulk_upload_template.xlsx"
    wb.save(filename)
    print(f"✅ Excel template created: {filename}")
    print(f"📊 Headers: {', '.join(headers)}")
    print(f"📝 Sample rows: {len(sample_data)}")
    print("\n📋 Column Descriptions:")
    print("  - entity_type: ASSET, COMPONENT, AMC, WARRANTY, CATEGORY, SUBCATEGORY, MAKE, MODEL, OUTLET, VENDOR")
    print("  - entity_id: ID of the entity (must exist in database)")
    print("  - file_name: Name of the file")
    print("  - file_path: Full path to file on server (file must already exist)")
    print("  - doc_type: PDF, IMAGE, RECEIPT, AGREEMENT (optional)")

if __name__ == "__main__":
    try:
        create_excel_template()
    except ImportError:
        print("❌ Error: openpyxl library not found")
        print("📦 Install it using: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error creating template: {e}")

